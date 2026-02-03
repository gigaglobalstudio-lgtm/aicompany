"""
Finance Excel Generator
Usage: python excel_generator.py --price 50000 --cost 15000 --shipping 3000 --ads 7000 --fee 10
"""

import sys
import os
import argparse
from datetime import datetime

# Fix encoding for Windows
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.chart import BarChart, Reference
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not installed. Run: pip install openpyxl")

def calculate_financials(price, cost, shipping, ads, fee_percent):
    """Calculate all financial metrics."""
    platform_fee = price * (fee_percent / 100)
    total_cost = cost + shipping + platform_fee + ads
    net_profit = price - total_cost
    margin = (net_profit / price) * 100 if price > 0 else 0
    roi = (net_profit / total_cost) * 100 if total_cost > 0 else 0

    return {
        'price': price,
        'cost': cost,
        'shipping': shipping,
        'platform_fee': platform_fee,
        'ads': ads,
        'total_cost': total_cost,
        'net_profit': net_profit,
        'margin': margin,
        'roi': roi
    }

def print_ascii_table(data):
    """Print ASCII P&L table."""
    print("\n" + "=" * 50)
    print("           P&L Statement (Profit & Loss)")
    print("=" * 50)
    print(f"| Revenue (Sales)       |    W{data['price']:>12,} |")
    print(f"| (-) COGS              |    W{data['cost']:>12,} |")
    print(f"| (-) Shipping          |    W{data['shipping']:>12,} |")
    print(f"| (-) Platform Fee      |    W{data['platform_fee']:>12,.0f} |")
    print(f"| (-) Ads               |    W{data['ads']:>12,} |")
    print("-" * 50)
    print(f"| Net Profit            |    W{data['net_profit']:>12,.0f} |")
    print(f"| Margin Rate           |      {data['margin']:>10.1f}% |")
    print(f"| ROI                   |      {data['roi']:>10.1f}% |")
    print("=" * 50)

    # Warning if negative margin
    if data['net_profit'] < 0:
        print("\n[WARNING] Negative margin! Price adjustment needed.")
    elif data['margin'] < 20:
        print("\n[NOTICE] Margin below 20%.")

def generate_excel(data, output_path):
    """Generate Excel report with multiple sheets."""
    if not EXCEL_AVAILABLE:
        print("Excel generation skipped - openpyxl not installed")
        return False

    wb = Workbook()

    # Sheet 1: P&L Summary
    ws1 = wb.active
    ws1.title = "P&L Summary"

    # Styling
    header_font = Font(bold=True, size=14)
    money_font = Font(size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")

    # Title
    ws1['A1'] = "손익계산서 (P&L Statement)"
    ws1['A1'].font = Font(bold=True, size=16)
    ws1['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

    # Headers
    ws1['A4'] = "항목"
    ws1['B4'] = "금액 (₩)"
    ws1['A4'].fill = header_fill
    ws1['B4'].fill = header_fill
    ws1['A4'].font = header_font_white
    ws1['B4'].font = header_font_white

    # Data
    rows = [
        ("매출액 (Revenue)", data['price']),
        ("(-) 원가 (COGS)", data['cost']),
        ("(-) 배송비 (Shipping)", data['shipping']),
        ("(-) 플랫폼수수료", data['platform_fee']),
        ("(-) 광고비 (Ads)", data['ads']),
        ("", ""),
        ("순이익 (Net Profit)", data['net_profit']),
        ("마진율 (%)", f"{data['margin']:.1f}%"),
        ("ROI (%)", f"{data['roi']:.1f}%"),
    ]

    for i, (label, value) in enumerate(rows, start=5):
        ws1[f'A{i}'] = label
        if isinstance(value, (int, float)) and label:
            ws1[f'B{i}'] = value
            ws1[f'B{i}'].number_format = '₩#,##0'
        else:
            ws1[f'B{i}'] = value

    # Adjust column widths
    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 20

    # Sheet 2: Scenario Analysis
    ws2 = wb.create_sheet("Scenario Analysis")
    ws2['A1'] = "시나리오 분석"
    ws2['A1'].font = Font(bold=True, size=16)

    scenarios = [
        ("Best Case", 1.2, 0.8),   # 120% sales, 80% ads
        ("Base Case", 1.0, 1.0),
        ("Worst Case", 0.6, 1.2),  # 60% sales, 120% ads
    ]

    ws2['A3'] = "시나리오"
    ws2['B3'] = "순이익"
    ws2['C3'] = "마진율"

    for i, (name, sales_mult, ads_mult) in enumerate(scenarios, start=4):
        adjusted_ads = data['ads'] * ads_mult
        adjusted_data = calculate_financials(
            data['price'], data['cost'], data['shipping'],
            adjusted_ads, (data['platform_fee'] / data['price']) * 100
        )
        ws2[f'A{i}'] = name
        ws2[f'B{i}'] = adjusted_data['net_profit'] * sales_mult
        ws2[f'B{i}'].number_format = '₩#,##0'
        ws2[f'C{i}'] = f"{adjusted_data['margin']:.1f}%"

    # Sheet 3: Monthly Projection
    ws3 = wb.create_sheet("Monthly Projection")
    ws3['A1'] = "12개월 수익 예측"
    ws3['A1'].font = Font(bold=True, size=16)

    ws3['A3'] = "월"
    ws3['B3'] = "예상 판매량"
    ws3['C3'] = "월 순이익"
    ws3['D3'] = "누적 순이익"

    cumulative = 0
    base_units = 100  # Starting with 100 units
    for month in range(1, 13):
        units = int(base_units * (1 + (month - 1) * 0.1))  # 10% growth
        monthly_profit = data['net_profit'] * units
        cumulative += monthly_profit

        row = month + 3
        ws3[f'A{row}'] = f"{month}월"
        ws3[f'B{row}'] = units
        ws3[f'C{row}'] = monthly_profit
        ws3[f'C{row}'].number_format = '₩#,##0'
        ws3[f'D{row}'] = cumulative
        ws3[f'D{row}'].number_format = '₩#,##0'

    # Save
    wb.save(output_path)
    return True

def main():
    parser = argparse.ArgumentParser(description='Finance Excel Generator')
    parser.add_argument('--price', type=float, required=True, help='Selling price')
    parser.add_argument('--cost', type=float, required=True, help='Product cost (COGS)')
    parser.add_argument('--shipping', type=float, default=0, help='Shipping cost')
    parser.add_argument('--ads', type=float, default=0, help='Advertising cost')
    parser.add_argument('--fee', type=float, default=10, help='Platform fee percentage')
    parser.add_argument('--output', type=str, default='Output/finance_report.xlsx', help='Output file path')

    args = parser.parse_args()

    # Calculate
    data = calculate_financials(args.price, args.cost, args.shipping, args.ads, args.fee)

    # Print ASCII table
    print_ascii_table(data)

    # Generate Excel
    output_path = args.output
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)

    if generate_excel(data, output_path):
        print(f"\n✅ Excel 저장 완료: {output_path}")

    return data

if __name__ == "__main__":
    main()
